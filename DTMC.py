from __future__ import annotations

import os
import io
import math
from datetime import datetime, date

import pandas as pd
import plotly.graph_objects as go
import streamlit as st

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "DTMC stats.xlsx"
)

APP_VERSION = "v25 — 2026-07-13"

REPORT_TITLE = "Financial Performance of RG Participants - FY 2025"

METRIC_COL = "In MM (USD)"

SCALE_NOTE = "Revenue and Net Income are in USD millions ($MM)."

DATA_SOURCE_NOTE = "Source: S&P Capital IQ"

CONSOLIDATED_NOTE = (
    "Merrill Lynch and Citibank NA reflect the consolidated financials of "
    "Bank of America and Citigroup, respectively."
)

# Region for each bank column. Banks not listed fall back to "Other" — add new
# banks here to place them in a region.
BANK_REGIONS = {
    "TD": "Canada",
    "RBC": "Canada",
    "BNS": "Canada",
    "BMO": "Canada",
    "CIBC": "Canada",
    "NBC": "Canada",
    "LBC": "Canada",
    "Desjardins": "Canada",
    "ATB": "Canada",
    "BNP Paribas": "Europe",
    "Merrill Lynch": "United States",
    "Citibank NA": "United States",
}


def bank_region(bank):
    return BANK_REGIONS.get(str(bank).strip(), "Other")


# Full legal/common names for the legend at the bottom of the page and in
# the PDF. Banks not listed fall back to their column name.
BANK_FULL_NAMES = {
    "TD": "Toronto-Dominion Bank",
    "RBC": "Royal Bank of Canada",
    "BNS": "Bank of Nova Scotia (Scotiabank)",
    "BMO": "Bank of Montreal",
    "CIBC": "Canadian Imperial Bank of Commerce",
    "NBC": "National Bank of Canada",
    "LBC": "Laurentian Bank of Canada",
    "Desjardins": "Desjardins Group",
    "ATB": "ATB Financial",
    "BNP Paribas": "BNP Paribas S.A.",
    "Merrill Lynch": "Merrill Lynch (Bank of America)",
    "Citibank NA": "Citibank N.A. (Citigroup)",
}


def full_bank_name(bank):
    return BANK_FULL_NAMES.get(str(bank).strip(), str(bank))


# Reporting-date overrides: these take precedence over the Date row in the
# Excel file. Remove an entry (or the whole dict) to fall back to the sheet.
REPORTING_DATE_OVERRIDES = {
    "TD": "October-2025",
    "RBC": "October-2025",
    "BNS": "October-2025",
    "BMO": "October-2025",
    "CIBC": "October-2025",
    "NBC": "October-2025",
    "LBC": "October-2025",
    "Desjardins": "December-2025",
    "BNP Paribas": "December-2025",
    "Merrill Lynch": "December-2025",
    "Citibank NA": "December-2025",
    "ATB": "March-2025",
}


# Short display names for chart x-axes (full name still shown on hover).
# Banks not listed fall back to their full name.
SHORT_NAMES = {
    "BNP Paribas": "BNP",
    "Merrill Lynch": "ML",
    "Citibank NA": "Citi",
    "Desjardins": "DESJ",
}


def short_name(bank):
    return SHORT_NAMES.get(str(bank).strip(), str(bank))


def compact_label(value, fmt):
    """On-bar labels. Currency shows the full $MM value; percents and plain
    numbers drop trailing zeros (142% not 142.00%, 4 not 4.00) so labels stay
    narrow and don't collide."""
    if fmt == "currency":
        return f"${value:,.0f}"
    if fmt == "percent":
        return f"{value:,.1f}%"
    return f"{value:,.1f}"


MISSING_TOKENS = {"", "na", "n/a", "n.a.", "-", "—", "nm", "nmf"}

# Rows with these names (case-insensitive) are treated as per-bank reporting
# dates, not metrics: pulled out of the charts/heatmap and shown as captions.
DATE_ROW_NAMES = {"date", "as of", "as-of", "as of date", "reporting date"}

PERFORMANCE_KEYWORDS = (
    "revenue", "income", "earnings", "profit", "margin",
    "yoy", "growth", "eps", "roe", "roa"
)

# Checked BEFORE the performance keywords, so "Equity Price" doesn't get
# claimed by another topic.
MARKET_KEYWORDS = (
    "cds", "equity price", "share price", "stock", "spread", "market"
)

TOPIC_PERFORMANCE = "Financial Performance"
TOPIC_RISK = "Capital, Liquidity & Asset Quality"
TOPIC_MARKET = "Market Indicators"

TOPIC_ICONS = {
    TOPIC_PERFORMANCE: "📈",
    TOPIC_RISK: "🛡️",
    TOPIC_MARKET: "💹",
}

# CDS charts invert the color logic: a POSITIVE change (spread widening) is
# BAD news for credit perception, a NEGATIVE change (tightening) is good.
MARKET_NOTE = (
    "Market indicators are as of June 2026 — the ▲ 3 months change is "
    "measured vs. March 2026, and the ▲ 1 year change vs. June 2025."
)

# (internal metric name, bank) pairs excluded from charts and the heatmap.
# "YoY (▲%) (2)" is the Net Income YoY row (second YoY row in the sheet).
CHART_EXCLUSIONS = {("YoY (▲%) (2)", "LBC")}


def chart_excluded_banks(metric):
    return [b for (m, b) in CHART_EXCLUSIONS if m == metric]


# Industry-median benchmarks. Each group lists the metrics it covers
# (matched case-insensitively against metric names) and the reference value
# for each. Add a group here to make it appear in the sidebar selector.
INDUSTRY_MEDIAN_GROUPS = {
    "Canadian banks + Merrill Lynch": {
        "cet 1": 12.9,
        "lcr": 127.0,
    },
    "Citigroup": {
        "cet 1": 15.0,
    },
    "BNP Paribas": {
        "cet 1": 17.4,
        "lcr": 181.0,
    },
}

# Colors used for the reference line, keyed to group name.
MEDIAN_GROUP_COLORS = {
    "Canadian banks + Merrill Lynch": "#c0392b",  # red
    "Citigroup":                      "#8a6d3b",  # brown
    "BNP Paribas":                    "#6b4b8a",  # purple
}


def _median_key_for(metric):
    """Return the config key inside a benchmark group that matches this
    metric's name, or None."""
    name = clean_metric_name(metric).lower()
    for group_data in INDUSTRY_MEDIAN_GROUPS.values():
        for key in group_data:
            if key in name:
                return key
    return None


def industry_median_for_group(metric, group):
    """Return (label, value) for `group`'s benchmark on this metric, or None."""
    if group not in INDUSTRY_MEDIAN_GROUPS:
        return None
    key = _median_key_for(metric)
    if key is None:
        return None
    value = INDUSTRY_MEDIAN_GROUPS[group].get(key)
    if value is None:
        return None
    display = f"{value:g}%" if key in ("cet 1", "lcr") else f"{value:g}"
    return (f"Industry median — {group}: {display}", value)


ACCENT = "#1f6f8b"   # neutral bars
NEG = "#c0392b"      # negative values only (positive for CDS widening)
POS = "#2a9d4a"      # CDS tightening only
HILITE = "#e08a1e"   # highlighted bank
GRID_CLR = "#eef2f4"
ZERO_CLR = "#c9d2d8"
STRIPE_BG = "#f8fbfc"

# Palette used to color each year's bars in a multi-year chart. The most
# recent year always gets the first (dark teal) color; older years step
# through the palette below.
YEAR_PALETTE = ["#1f6f8b", "#c98a3e", "#8a6fae", "#5b8c5a", "#c0392b"]


def year_color(year, years_in_family):
    """Stable color for `year` within a family, most-recent year first."""
    years_sorted = sorted(
        (y for y in years_in_family if y is not None), reverse=True
    )
    if year not in years_sorted:
        return ACCENT
    return YEAR_PALETTE[years_sorted.index(year) % len(YEAR_PALETTE)]


def make_unique_index(index):
    counts = {}
    new_index = []

    for item in index:
        item = str(item).strip()

        if item not in counts:
            counts[item] = 1
            new_index.append(item)
        else:
            counts[item] += 1
            new_index.append(f"{item} ({counts[item]})")

    return new_index


def _is_year_suffix(digits):
    """True if a bracketed numeric suffix like '2025' looks like a year
    rather than a de-dupe counter (the '(2)', '(3)' that make_unique_index
    appends to repeated metric names, e.g. the two 'YoY (▲%)' rows)."""
    return len(digits) == 4 and digits.isdigit() and 1900 <= int(digits) <= 2100


def clean_metric_name(metric):
    """Strip a trailing de-dupe counter like ' (2)' from a metric name.
    Does NOT strip a trailing 4-digit year like ' (2025)' — that's real
    metadata, not a de-dupe artifact."""
    metric = str(metric)

    if metric.endswith(")") and " (" in metric:
        base, suffix = metric.rsplit(" (", 1)
        digits = suffix[:-1]
        if digits.isdigit() and not _is_year_suffix(digits):
            return base

    return metric


def parse_year_suffix(metric):
    """Split 'Revenue (2025)' into ('Revenue', 2025). If there's no trailing
    4-digit year, returns (metric, None) unchanged."""
    s = str(metric)

    if s.endswith(")") and " (" in s:
        base, suffix = s.rsplit(" (", 1)
        digits = suffix[:-1]
        if _is_year_suffix(digits):
            return base, int(digits)

    return s, None


def group_into_families(metrics):
    """Group metric rows that share a base name (ignoring the year suffix)
    into "families" — e.g. 'Revenue (2025)' and 'Revenue (2024)' become one
    family with two (year, metric) entries. Metrics without a year suffix
    become a family of one. Returns [(family_base_name, [(year, metric), ...])]
    preserving the order families first appear."""
    families = {}
    order = []

    for m in metrics:
        base, year = parse_year_suffix(m)
        if base not in families:
            families[base] = []
            order.append(base)
        families[base].append((year, m))

    return [(base, families[base]) for base in order]


def resolve_family_bank_order(entries, num_v, sort_mode):
    """Decide which banks appear on a family's chart, and in what order,
    based on the most-recent year's values. Honors the three sort modes."""

    def usable_series(metric):
        s = num_v.loc[metric].dropna()
        excluded = chart_excluded_banks(metric)
        return s.drop(index=[b for b in excluded if b in s.index])

    years = sorted((y for y, _ in entries if y is not None), reverse=True)
    if years:
        baseline_metric = next(m for y, m in entries if y == years[0])
    else:
        baseline_metric = entries[0][1]

    baseline_series = usable_series(baseline_metric)

    all_banks = []
    for _, m in entries:
        for b in usable_series(m).index:
            if b not in all_banks:
                all_banks.append(b)

    if sort_mode == "Alphabetical":
        all_banks.sort(key=lambda b: str(b).lower())
    elif sort_mode == "By value":
        all_banks.sort(
            key=lambda b: -(baseline_series[b] if b in baseline_series.index
                            else float("-inf"))
        )
    elif sort_mode == "By region, then value":
        all_banks.sort(
            key=lambda b: (bank_region(b),
                           -(baseline_series[b] if b in baseline_series.index
                             else float("-inf")))
        )
    else:  # "File order"
        col_order = list(num_v.columns)
        all_banks.sort(
            key=lambda b: col_order.index(b) if b in col_order else 999
        )

    return all_banks


def is_cds_metric(metric):
    return "cds" in clean_metric_name(metric).lower()


def display_metric_name(metric):
    """Chart/table title: cleaned name with '%' markers removed —
    'Equity Price %(▲ 3 months)' → 'Equity Price (▲ 3 months)',
    'Gross NPAs/... + OREO (%)' → 'Gross NPAs/... + OREO',
    'YoY (▲%)' → 'YoY (▲)'."""
    s = clean_metric_name(metric)
    s = s.replace("(▲%)", "(▲)")
    s = s.replace("%(", "(")
    s = s.replace("(%)", "")
    return " ".join(s.split())


def display_labels(metrics):
    """Cleaned metric names made unique again for pandas (Styler requires a
    unique index). Duplicates get zero-width spaces appended, so 'YoY (▲%)'
    can appear twice looking identical while remaining distinct labels."""
    seen = {}
    labels = []
    for m in metrics:
        name = display_metric_name(m)
        seen[name] = seen.get(name, 0) + 1
        labels.append(name + "\u200b" * (seen[name] - 1))
    return labels


def parse_value(raw):
    if raw is None:
        return None

    try:
        if pd.isna(raw):
            return None
    except Exception:
        pass

    s = str(raw).strip()

    if s.lower() in MISSING_TOKENS:
        return None

    negative = False

    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1]

    s = s.replace("$", "").replace(",", "").replace("%", "").strip()

    if s.startswith("-"):
        negative = True
        s = s[1:]

    try:
        value = float(s)
    except ValueError:
        return None

    return -value if negative else value


def detect_format(metric_name, raw_values):
    name = clean_metric_name(metric_name).lower()
    cells = [str(v) for v in raw_values if v is not None]

    currency_keywords = ["revenue", "income"]
    percent_keywords = [
        "%", "yoy", "ratio", "lcr", "npas", "loans",
        "equity price", "cet", "roe", "roa"
    ]

    if any(k in name for k in currency_keywords):
        return "currency"

    if any(k in name for k in percent_keywords) or any("%" in c for c in cells):
        return "percent"

    if any("$" in c for c in cells):
        return "currency"

    return "number"


def format_value(value, fmt, original=""):
    try:
        if pd.isna(value):
            return str(original) if original not in (None, "") else "—"
    except Exception:
        return str(original)

    try:
        value = float(value)
    except Exception:
        return str(original)

    if fmt == "currency":
        return f"${value:,.0f}"

    if fmt == "percent":
        # Values arrive already scaled (the loader converts Excel's stored
        # fractions using each cell's display format), so no ×100 guessing.
        return f"{value:,.1f}%"

    return f"{value:,.1f}"


def group_metrics_by_topic(metrics, formats):
    perf, risk, market = [], [], []

    for m in metrics:
        name = clean_metric_name(m).lower()

        if any(k in name for k in MARKET_KEYWORDS):
            market.append(m)
        elif formats.get(m) == "currency" or any(k in name for k in PERFORMANCE_KEYWORDS):
            perf.append(m)
        else:
            risk.append(m)

    # Fallback: if everything landed in a single bucket, split evenly across
    # the first two topics so the pages are still useful.
    buckets = [perf, risk, market]
    if metrics and sum(1 for b in buckets if b) == 1:
        half = math.ceil(len(metrics) / 2)
        perf, risk, market = list(metrics[:half]), list(metrics[half:]), []

    return [
        (TOPIC_PERFORMANCE, perf),
        (TOPIC_RISK, risk),
        (TOPIC_MARKET, market),
    ]


def load_raw(source):
    """Load the Excel sheet using each cell's DISPLAY format, so a cell stored
    as 1.42 with a percent format arrives as '142%', 0.0054 as '0.54%', and
    59180 with a '$' format as '$59,180'. This fixes the LCR scaling issue and
    removes the need for any ×100 heuristics downstream. Dates are rendered
    per their format ('mmm-yy' → 'Apr-26', 'dd-mmm-yy' → '01-Mar-26'). Text
    like 'Meets Req' or 'NA' passes through verbatim."""
    from openpyxl import load_workbook

    wb = load_workbook(source, data_only=True)
    ws = wb.active

    def cell_to_str(cell):
        v = cell.value

        if v is None:
            return ""

        if isinstance(v, str):
            return v.strip()

        fmt = (cell.number_format or "").lower()

        if isinstance(v, (datetime, date)):
            if "d" in fmt:
                return v.strftime("%d-%b-%y")
            return v.strftime("%b-%y")

        if "%" in fmt:
            return f"{round(v * 100, 4):g}%"

        if "$" in fmt:
            return f"${v:,.0f}"

        if isinstance(v, float) and v == int(v):
            return str(int(v))

        return str(v)

    rows = [[cell_to_str(c) for c in row] for row in ws.iter_rows()]

    if not rows:
        return pd.DataFrame(columns=[METRIC_COL]).set_index(METRIC_COL)

    header = [h if h else f"Column {i}" for i, h in enumerate(rows[0])]
    header[0] = METRIC_COL

    body = [r for r in rows[1:] if r and str(r[0]).strip()]

    df = pd.DataFrame(body, columns=header)
    df = df.set_index(METRIC_COL)
    df.index = make_unique_index(df.index)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.loc[:, [c for c in df.columns if c and not c.startswith("Column ")]]

    return df


def split_date_row(raw):
    """Pull the per-bank reporting-date row (e.g. 'Date') out of the metrics.

    Returns (metrics_frame, date_series_or_None)."""
    for m in raw.index:
        if clean_metric_name(m).strip().lower() in DATE_ROW_NAMES:
            dates = raw.loc[m]
            return raw.drop(index=m), dates
    return raw, None


def long_date_label(date_str):
    """'Apr-26' or '01-Mar-26' → 'April-2026'. Unparseable text passes
    through as-is."""
    s = str(date_str).strip()

    if not s or s == "—":
        return "—"

    d = None
    for fmt in ("%b-%y", "%d-%b-%y"):
        try:
            d = pd.to_datetime(s, format=fmt)
            break
        except Exception:
            continue

    if d is None:
        try:
            d = pd.to_datetime(s, dayfirst=True)
        except Exception:
            return s

    return f"{d.strftime('%B')}-{d.year}"


def date_caption_lines(dates, banks):
    """One line per reporting date, grouping the banks that share it:
    ['April-2026 — TD, RBC, …', 'December-2025 — Desjardins', …]"""
    if dates is None:
        return []

    groups = {}
    for b in banks:
        d = long_date_label(dates.get(b, ""))
        groups.setdefault(d, []).append(b)

    return [f"{d} — {', '.join(bs)}" for d, bs in groups.items()]


def build_numeric(raw):
    formats = {}
    numeric_rows = []

    for metric, row in raw.iterrows():
        formats[metric] = detect_format(metric, row.tolist())
        numeric_rows.append([parse_value(v) for v in row])

    numeric = pd.DataFrame(
        numeric_rows,
        index=raw.index,
        columns=raw.columns
    )

    return numeric.astype("float64"), formats


# --------------------------------------------------------------------------- #
# PDF report (reportlab for layout, matplotlib for static charts)
# --------------------------------------------------------------------------- #
def _metric_png(metric, num_v, fmt, sort_mode, highlight, median_group):
    """Render one metric's bar chart to PNG bytes with matplotlib (Agg),
    mirroring the on-screen chart: same colors, ordering, and highlight."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.ticker import FuncFormatter

    series = num_v.loc[metric].dropna()
    series = series.drop(index=[b for b in chart_excluded_banks(metric)
                                if b in series.index])

    if series.empty:
        return None

    if sort_mode == "Alphabetical":
        order = sorted(series.index, key=lambda b: str(b).lower())
        series = series.loc[order]
    elif sort_mode == "By value":
        series = series.sort_values(ascending=False)
    elif sort_mode == "By region, then value":
        order = sorted(series.index,
                       key=lambda b: (bank_region(b), -series[b]))
        series = series.loc[order]

    banks = list(series.index)
    values = series.values

    if is_cds_metric(metric):
        # Inverted: widening (positive) = red, tightening (negative) = green.
        bar_colors = [NEG if v > 0 else (POS if v < 0 else ACCENT)
                      for v in values]
    else:
        bar_colors = [NEG if v < 0 else ACCENT for v in values]

    if highlight != "(none)":
        bar_colors = [
            HILITE if b == highlight else c
            for b, c in zip(banks, bar_colors)
        ]

    fig, ax = plt.subplots(figsize=(6.4, 3.4), dpi=150)
    bars = ax.bar([short_name(b) for b in banks], values, color=bar_colors)
    ax.grid(axis="y", color=GRID_CLR, linewidth=0.6)
    ax.set_axisbelow(True)
    ax.margins(y=0.18)  # headroom for the staggered labels

    if median_group:
        median = industry_median_for_group(metric, median_group)
        if median:
            label, value = median
            clr = MEDIAN_GROUP_COLORS.get(median_group, "#5b6770")
            ax.axhline(value, color=clr, linewidth=1.4, linestyle="--")
            ax.annotate(label, xy=(0.5, value),
                        xycoords=("axes fraction", "data"),
                        fontsize=7, color=clr, ha="center", va="bottom",
                        fontweight="bold",
                        bbox=dict(boxstyle="round,pad=0.2",
                                  fc="white", ec=clr, lw=0.6))

    ax.set_title(display_metric_name(metric), fontsize=10, fontweight="bold")
    ax.axhline(0, color="#888888", linewidth=0.6)
    ax.spines[["top", "right"]].set_visible(False)
    ax.tick_params(axis="x", labelrotation=45, labelsize=7)
    for lbl in ax.get_xticklabels():
        lbl.set_ha("right")
    ax.tick_params(axis="y", labelsize=7)

    if fmt == "currency":
        ax.set_ylabel("US$ MM", fontsize=7, color="#666666")
        ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: f"${v:,.0f}"))
        fmt_key = "currency"
    elif fmt == "percent":
        ax.yaxis.set_major_formatter(
            FuncFormatter(lambda v, _: f"{v:,.1f}%"))
        fmt_key = "percent"
    else:
        ax.yaxis.set_major_formatter(
            FuncFormatter(lambda v, _: f"{v:,.1f}"))
        fmt_key = "number"

    # On-bar numeric labels are intentionally omitted to keep the chart
    # legible with many banks; exact values are available in the table.

    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    plt.close(fig)
    return buf.getvalue()


def _family_png_multi(family_base, entries, num_v, fmt, sort_mode,
                      highlight, median_group):
    """Render a multi-year family (e.g. Revenue 2025 + 2024) as one grouped
    bar chart: one bar cluster per bank, one color per year, legend at the
    bottom of the chart."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np
    from matplotlib.ticker import FuncFormatter

    entries = sorted(entries, key=lambda e: (e[0] is None, e[0] or 0))
    years = [y for y, _ in entries]
    banks = resolve_family_bank_order(entries, num_v, sort_mode)

    if not banks:
        return None

    n_years = len(entries)
    x = np.arange(len(banks))
    width = 0.8 / n_years

    fig, ax = plt.subplots(figsize=(6.4, 3.4), dpi=150)

    for i, (year, metric) in enumerate(entries):
        excluded = set(chart_excluded_banks(metric))
        series = num_v.loc[metric]
        values = [
            series[b] if (b in series.index and b not in excluded
                          and pd.notna(series[b])) else float("nan")
            for b in banks
        ]
        offset = (i - (n_years - 1) / 2) * width
        bars = ax.bar(
            x + offset, values, width=width,
            color=year_color(year, years),
            label=str(year) if year is not None else display_metric_name(metric),
        )
        if highlight != "(none)" and highlight in banks:
            bars[banks.index(highlight)].set_edgecolor(HILITE)
            bars[banks.index(highlight)].set_linewidth(1.6)

    ax.set_xticks(x)
    ax.set_xticklabels([short_name(b) for b in banks], rotation=45,
                       ha="right", fontsize=7)
    ax.grid(axis="y", color=GRID_CLR, linewidth=0.6)
    ax.set_axisbelow(True)
    ax.margins(y=0.18)
    ax.axhline(0, color="#888888", linewidth=0.6)
    ax.spines[["top", "right"]].set_visible(False)
    ax.tick_params(axis="y", labelsize=7)
    ax.set_title(display_metric_name(family_base), fontsize=10,
                 fontweight="bold")

    if fmt == "currency":
        ax.set_ylabel("US$ MM", fontsize=7, color="#666666")
        ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: f"${v:,.0f}"))
    elif fmt == "percent":
        ax.yaxis.set_major_formatter(
            FuncFormatter(lambda v, _: f"{v:,.1f}%"))
    else:
        ax.yaxis.set_major_formatter(
            FuncFormatter(lambda v, _: f"{v:,.1f}"))

    if median_group:
        median = industry_median_for_group(family_base, median_group)
        if median:
            label, value = median
            clr = MEDIAN_GROUP_COLORS.get(median_group, "#5b6770")
            ax.axhline(value, color=clr, linewidth=1.4, linestyle="--")
            ax.annotate(label, xy=(0.5, value),
                        xycoords=("axes fraction", "data"),
                        fontsize=7, color=clr, ha="center", va="bottom",
                        fontweight="bold",
                        bbox=dict(boxstyle="round,pad=0.2",
                                  fc="white", ec=clr, lw=0.6))

    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.32),
              ncol=n_years, fontsize=7, frameon=False)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    plt.close(fig)
    return buf.getvalue()


def _family_png(family_base, entries, num_v, formats, sort_mode, highlight,
                median_group):
    """Dispatch a family to the single-year or multi-year PNG renderer."""
    fmt = formats[entries[0][1]]
    if len(entries) <= 1:
        return _metric_png(entries[0][1], num_v, fmt, sort_mode, highlight,
                           median_group)
    return _family_png_multi(family_base, entries, num_v, fmt, sort_mode,
                             highlight, median_group)


def build_pdf_report(raw_v, num_v, formats, source_label,
                     sort_mode, highlight, dates_items, median_group,
                     use_option_1):
    """Assemble a landscape PDF: title + comparison table (with 'Region' and
    'As of' columns) on page 1, then charts grouped by topic -- each topic
    starts on its own landscape page."""
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, Image, PageBreak)

    metrics = list(raw_v.index)
    banks = list(raw_v.columns)
    dates = dict(dates_items) if dates_items else {}
    page_w, page_h = landscape(letter)
    margin = 0.5 * inch
    avail_w = page_w - 2 * margin

    styles = getSampleStyleSheet()
    h_cell = ParagraphStyle("hcell", parent=styles["Normal"], fontSize=6.5,
                            leading=8, textColor=colors.white,
                            fontName="Helvetica-Bold")
    row_lbl = ParagraphStyle("rowlbl", parent=styles["Normal"], fontSize=7,
                             leading=8, fontName="Helvetica-Bold")
    subtitle = ParagraphStyle("sub", parent=styles["Normal"], fontSize=8,
                              textColor=colors.HexColor("#666666"))

    story = []

    # --- Comparison table (alphabetical, no market indicators). --------------
    alpha_banks = sorted(banks, key=lambda b: str(b).lower())
    table_metrics = [
        m for m in metrics
        if not any(k in clean_metric_name(m).lower() for k in MARKET_KEYWORDS)
    ]

    header = [Paragraph("Bank", h_cell)]
    header += [Paragraph(display_metric_name(m), h_cell) for m in table_metrics]
    table_data = [header]
    red_cells = []
    for r, bank in enumerate(alpha_banks, start=1):
        cells = [Paragraph(bank, row_lbl)]
        for c, metric in enumerate(table_metrics, start=1):
            val = num_v.loc[metric, bank]
            cells.append(format_value(val, formats[metric],
                                      raw_v.loc[metric, bank]))
            if pd.notna(val) and val < 0:
                red_cells.append((c, r))
        table_data.append(cells)

    first_w = 1.0 * inch
    other_w = (avail_w - first_w) / max(len(table_metrics), 1)
    tbl = Table(table_data,
                colWidths=[first_w] + [other_w] * len(table_metrics),
                repeatRows=1)
    tbl_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(ACCENT)),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#f3f7f9")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for c, r in red_cells:
        tbl_style.append(("TEXTCOLOR", (c, r), (c, r), colors.HexColor(NEG)))
    tbl.setStyle(TableStyle(tbl_style))
    story.append(tbl)

    # Bank legend + notes directly below the table.
    legend = ";  ".join(f"<b>{b}</b> = {full_bank_name(b)}" for b in alpha_banks)
    story += [
        Spacer(1, 6),
        Paragraph(legend, subtitle),
        Spacer(1, 4),
        Paragraph(SCALE_NOTE, subtitle),
        Paragraph(CONSOLIDATED_NOTE, subtitle),
    ]

    # Reusable Industry medians line (shown on the Financial Performance page).
    _median_lines = []
    for _group_name, _group_data in INDUSTRY_MEDIAN_GROUPS.items():
        _parts = []
        for _key, _value in _group_data.items():
            _dm = "CET 1 ratio" if _key == "cet 1" else _key.upper()
            _parts.append(f"{_dm} {_value:g}%")
        _median_lines.append(f"<b>{_group_name}</b> — " + ", ".join(_parts))
    _industry_medians_note = ("Industry medians: "
                              + " &nbsp;·&nbsp; ".join(_median_lines))
    img_w = (avail_w - 0.2 * inch) / 2
    img_h = img_w * 0.54  # keeps 2 chart rows + notes on one landscape page

    for topic, topic_metrics in group_metrics_by_topic(metrics, formats):
        if not topic_metrics:
            continue

        # Every topic starts on a new page (the table sits on page 1).
        story.append(PageBreak())
        story.append(Paragraph(topic, styles["Heading2"]))

        if topic == TOPIC_MARKET:
            story.append(Paragraph(MARKET_NOTE, subtitle))
        else:
            for line in date_caption_lines(
                    pd.Series(dates) if dates else None, banks):
                story.append(Paragraph(line, subtitle))
            if topic == TOPIC_PERFORMANCE:
                story.append(Paragraph(_industry_medians_note, subtitle))
        story.append(Spacer(1, 4))

        chart_rows = []
        if use_option_1 and topic == TOPIC_PERFORMANCE:
            # Pair each currency family with the next YoY family: currency
            # LEFT, YoY RIGHT, one row each.
            left_png = None
            for family_base, entries in group_into_families(topic_metrics):
                png = _family_png(family_base, entries, num_v, formats,
                                  sort_mode, highlight, median_group)
                if png is None:
                    continue
                is_yoy = "yoy" in family_base.lower()
                img = Image(io.BytesIO(png), width=img_w, height=img_h)
                if is_yoy:
                    chart_rows.append([left_png or "", img])
                    left_png = None
                else:
                    if left_png is not None:
                        chart_rows.append([left_png, ""])
                    left_png = img
            if left_png is not None:
                chart_rows.append([left_png, ""])
        else:
            pair = []
            for family_base, entries in group_into_families(topic_metrics):
                png = _family_png(family_base, entries, num_v, formats,
                                  sort_mode, highlight, median_group)
                if png is None:
                    continue
                pair.append(Image(io.BytesIO(png),
                                  width=img_w, height=img_h))
                if len(pair) == 2:
                    chart_rows.append(pair)
                    pair = []
            if pair:
                pair.append("")
                chart_rows.append(pair)
        if chart_rows:
            grid = Table(chart_rows, colWidths=[img_w + 0.1 * inch] * 2)
            grid.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(grid)

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#999999"))
        canvas.drawRightString(page_w - margin, 0.3 * inch, f"Page {doc.page}")
        canvas.drawString(margin, 0.3 * inch,
                          f"{REPORT_TITLE} — Revenue & Net Income in MM (USD)")
        canvas.drawCentredString(page_w / 2, 0.3 * inch, DATA_SOURCE_NOTE)
        canvas.restoreState()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                            leftMargin=margin, rightMargin=margin,
                            topMargin=margin, bottomMargin=0.6 * inch,
                            title=REPORT_TITLE)
    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buf.getvalue()


@st.cache_data(show_spinner="Building PDF report…")
def get_report_bytes(raw_csv, formats_items, metrics, banks, source_label,
                     sort_mode, highlight, dates_items, median_group,
                     use_option_1, app_version=APP_VERSION):
    """Cached wrapper so the PDF only rebuilds when the data, selection,
    sort order, or highlight changes."""
    raw_v = pd.read_csv(io.StringIO(raw_csv), dtype=str,
                        keep_default_na=False).fillna("")
    raw_v = raw_v.set_index(raw_v.columns[0])
    raw_v = raw_v.loc[list(metrics), list(banks)]

    formats = dict(formats_items)

    numeric_rows = [[parse_value(v) for v in raw_v.loc[m]] for m in raw_v.index]
    num_v = pd.DataFrame(numeric_rows, index=raw_v.index,
                         columns=raw_v.columns).astype("float64")

    return build_pdf_report(raw_v, num_v, formats, source_label,
                            sort_mode, highlight, dates_items, median_group,
                            use_option_1)


st.set_page_config(
    page_title=REPORT_TITLE,
    page_icon="🏦",
    layout="wide",
)

st.title(f"🏦 {REPORT_TITLE}")

st.caption(f"{SCALE_NOTE} {CONSOLIDATED_NOTE}  \n{APP_VERSION}")

with st.sidebar:
    st.header("Data")

    source_choice = st.radio(
        "Source",
        ["Bundled Excel file", "Upload Excel file"],
        index=0,
    )

    upload = None

    if source_choice == "Upload Excel file":
        upload = st.file_uploader(
            "Excel file: first column = metrics, other columns = banks",
            type=["xlsx"],
        )

    if st.button("🔄 Reload data", use_container_width=True):
        st.rerun()


try:
    if upload is not None:
        raw = load_raw(io.BytesIO(upload.getvalue()))
        source_label = upload.name
    else:
        raw = load_raw(EXCEL_PATH)
        source_label = os.path.basename(EXCEL_PATH)

except FileNotFoundError:
    st.error(
        "Could not find `DTMC stats.xlsx`. Make sure it is in the same GitHub repository folder as `DTMC.py`."
    )
    st.stop()

except Exception as exc:
    st.error(f"Could not read the Excel file: {exc}")
    st.stop()


if raw.empty or raw.shape[1] == 0:
    st.warning("The Excel file has no bank columns.")
    st.stop()


# Pull the per-bank reporting-date row out of the metrics.
raw, report_dates = split_date_row(raw)

# Apply the code-level reporting-date overrides.
if REPORTING_DATE_OVERRIDES:
    if report_dates is None:
        report_dates = pd.Series("", index=raw.columns, dtype=object)
    for b, d in REPORTING_DATE_OVERRIDES.items():
        if b in report_dates.index:
            report_dates[b] = d

numeric, formats = build_numeric(raw)

all_banks = list(raw.columns)
all_metrics = list(raw.index)

with st.sidebar:
    st.header("Filters")

    all_regions = sorted({bank_region(b) for b in all_banks})

    sel_regions = st.multiselect(
        "Regions",
        all_regions,
        default=all_regions,
    )

    region_banks = [b for b in all_banks if bank_region(b) in sel_regions]

    sel_banks = st.multiselect(
        "Banks",
        region_banks,
        default=region_banks,
    )

    sel_metrics = st.multiselect(
        "Metrics",
        all_metrics,
        default=all_metrics,
    )

    st.divider()

    layout_option = st.radio(
        "Chart layout",
        ["Option 1 — Revenue & NI on left, YoY on right (alphabetical)",
         "Option 2 — Original (choose sort order below)"],
        index=0,
        help="Option 1 pairs each currency chart with its YoY chart in the "
             "same row and forces a shared alphabetical bank order, so bars "
             "line up across charts. Option 2 keeps the original layout "
             "with per-chart sorting.",
    )
    use_option_1 = layout_option.startswith("Option 1")

    if use_option_1:
        sort_mode = "Alphabetical"
        st.caption("Sort order: **Alphabetical** (locked in Option 1).")
    else:
        sort_mode = st.radio(
            "Bank order in charts",
            ["By value", "By region, then value", "Alphabetical", "File order"],
            index=0,
        )

    highlight = st.selectbox(
        "Highlight a bank",
        ["(none)"] + sel_banks,
        index=0,
    )

    median_group = st.radio(
        "Industry median",
        ["Off"] + list(INDUSTRY_MEDIAN_GROUPS.keys()),
        index=0,
        help="Overlay one benchmark group's median as a reference line on "
             "the CET1 and LCR charts.",
    )
    if median_group == "Off":
        median_group = None


if not sel_banks or not sel_metrics:
    st.info("Pick at least one bank and one metric.")
    st.stop()


raw_v = raw.loc[sel_metrics, sel_banks]
num_v = numeric.loc[sel_metrics, sel_banks]

# Reporting dates shown under each tab: one clear line per date group.
_date_lines = date_caption_lines(report_dates, sel_banks)
as_of_md = None
if _date_lines:
    as_of_md = "**Reporting date**  \n" + "  \n".join(
        f"**{line.split(' — ')[0]}** — {line.split(' — ', 1)[1]}"
        for line in _date_lines
    )


def show_reporting_dates():
    if as_of_md:
        with st.container(border=True):
            st.markdown(as_of_md)


def show_source_note():
    st.caption(DATA_SOURCE_NOTE)

with st.sidebar:
    st.divider()
    st.header("Report")
    st.caption(
        "Exports the table and all charts (grouped by topic, one landscape "
        "page each) for the currently selected banks and metrics, using the "
        "current sort order and highlight."
    )

    try:
        report_bytes = get_report_bytes(
            raw_v.reset_index().to_csv(index=False),
            tuple((m, formats[m]) for m in sel_metrics),
            tuple(sel_metrics),
            tuple(sel_banks),
            source_label,
            sort_mode,
            highlight,
            tuple((b, str(report_dates[b])) for b in sel_banks)
            if report_dates is not None else (),
            median_group,
            use_option_1,
            APP_VERSION,
        )

        st.download_button(
            "📄 Download report (PDF)",
            data=report_bytes,
            file_name=f"dtmc_stats_report_{datetime.now():%Y%m%d}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )

    except ModuleNotFoundError as exc:
        st.warning(
            f"PDF export needs an extra package: `{exc.name}`. "
            "Install with `pip install reportlab matplotlib` "
            "(add both to requirements.txt if deploying)."
        )

    except Exception as exc:
        st.error(f"Could not build the PDF: {exc}")

tab_charts, tab_table, tab_heat = st.tabs([
    "📊 Charts",
    "📋 Table",
    "🌡 Heatmap",
])


with tab_table:
    show_reporting_dates()

    display = pd.DataFrame(
        index=display_labels(raw_v.index),
        columns=raw_v.columns,
        dtype=object,
    )

    for i, m in enumerate(raw_v.index):
        for j, b in enumerate(raw_v.columns):
            display.iloc[i, j] = format_value(
                num_v.iloc[i, j],
                formats[m],
                raw_v.iloc[i, j],
            )

    # Keep the reporting dates visible as the last row of the table.
    if report_dates is not None:
        display.loc["Date"] = [long_date_label(report_dates[b])
                               for b in raw_v.columns]

    neg_mask = num_v.lt(0).values

    def _style(_):
        css = pd.DataFrame("", index=display.index, columns=display.columns)
        # Zebra striping to match the theme.
        for i in range(len(display.index)):
            if i % 2 == 1:
                css.iloc[i, :] = f"background-color: {STRIPE_BG};"
        # Negative values in red.
        for i in range(neg_mask.shape[0]):
            for j in range(neg_mask.shape[1]):
                if neg_mask[i, j]:
                    css.iloc[i, j] += f" color: {NEG}; font-weight: 600;"
        # Date row in muted italics.
        if report_dates is not None:
            css.iloc[-1, :] += " color: #6b7680; font-style: italic;"
        return css

    st.dataframe(
        display.style.apply(_style, axis=None),
        use_container_width=True,
    )

    st.caption(f"{SCALE_NOTE} {CONSOLIDATED_NOTE}")
    show_source_note()


with tab_heat:
    show_reporting_dates()

    norm = num_v.copy()
    for m, b in CHART_EXCLUSIONS:
        if m in norm.index and b in norm.columns:
            norm.loc[m, b] = float("nan")

    for i, _ in enumerate(norm.index):
        row = norm.iloc[i]
        lo = row.min()
        hi = row.max()

        if pd.isna(lo) or pd.isna(hi) or hi == lo:
            norm.iloc[i] = 0.5
        else:
            norm.iloc[i] = (row - lo) / (hi - lo)

    heat = go.Figure(
        go.Heatmap(
            z=norm.values,
            x=list(norm.columns),
            y=[display_metric_name(m) for m in norm.index],
            colorscale="Teal",
            zmin=0,
            zmax=1,
            xgap=2,
            ygap=2,
            hovertemplate="%{y}<br>%{x}<br>rank score: %{z:.1f}<extra></extra>",
            colorbar=dict(title="relative", thickness=12),
        )
    )

    heat.update_layout(
        height=60 + 42 * len(norm.index),
        margin=dict(l=10, r=10, t=10, b=10),
        yaxis=dict(autorange="reversed"),
        plot_bgcolor="rgba(0,0,0,0)",
    )

    st.plotly_chart(heat, use_container_width=True)
    show_source_note()


def _metric_bar_fig(metric, fmt):
    row_position = list(num_v.index).index(metric)
    series = num_v.iloc[row_position].dropna()
    series = series.drop(index=[b for b in chart_excluded_banks(metric)
                                if b in series.index])

    if series.empty:
        return None

    frame = series.reset_index()
    frame.columns = ["Bank", "Value"]
    frame["Region"] = [bank_region(b) for b in frame["Bank"]]

    if sort_mode == "Alphabetical":
        frame = frame.sort_values("Bank",
                                  key=lambda s: s.str.lower())
    elif sort_mode == "By value":
        frame = frame.sort_values("Value", ascending=False)
    elif sort_mode == "By region, then value":
        frame = frame.sort_values(["Region", "Value"],
                                  ascending=[True, False])
    # "File order": keep as-is

    if is_cds_metric(metric):
        # Inverted: widening (positive) = red, tightening (negative) = green.
        colors = [NEG if v > 0 else (POS if v < 0 else ACCENT)
                  for v in frame["Value"]]
    else:
        colors = [NEG if v < 0 else ACCENT for v in frame["Value"]]

    if highlight != "(none)":
        colors = [
            HILITE if b == highlight else c
            for b, c in zip(frame["Bank"], colors)
        ]

    # Title: metric name, plus a small scale note for currency metrics.
    subline_parts = []
    if fmt == "currency":
        subline_parts.append("in $ millions (USD)")

    title_text = display_metric_name(metric)
    if subline_parts:
        title_text += ("<br><sup style='color:#8a949c'>"
                       + "  ·  ".join(subline_parts) + "</sup>")

    if fmt == "currency":
        hovertmpl = ("<b>%{customdata[0]}</b> · %{customdata[1]}"
                     "<br>$%{y:,.0f}<extra></extra>")
        axfmt = "$,.0f"
    elif fmt == "percent":
        hovertmpl = ("<b>%{customdata[0]}</b> · %{customdata[1]}"
                     "<br>%{y:.1f}%<extra></extra>")
        axfmt = ".1f"
    else:
        hovertmpl = ("<b>%{customdata[0]}</b> · %{customdata[1]}"
                     "<br>%{y:,.1f}<extra></extra>")
        axfmt = ",.1f"

    fig = go.Figure(
        go.Bar(
            x=[short_name(b) for b in frame["Bank"]],
            y=frame["Value"],
            marker_color=colors,
            customdata=list(zip(frame["Bank"], frame["Region"])),
            hovertemplate=hovertmpl,
            cliponaxis=False,
        )
    )

    if median_group:
        median = industry_median_for_group(metric, median_group)
        if median:
            label, value = median
            clr = MEDIAN_GROUP_COLORS.get(median_group, "#5b6770")
            fig.add_hline(
                y=value, line_dash="dash", line_color=clr, line_width=2,
                annotation_text=f"<b>{label}</b>",
                annotation_position="top left",
                annotation_bgcolor="white",
                annotation_bordercolor=clr,
                annotation_borderwidth=1,
                annotation_borderpad=4,
                annotation_font=dict(size=12, color=clr),
            )

    fig.update_layout(
        title=dict(text=title_text, font=dict(size=14)),
        height=300,
        margin=dict(l=10, r=10, t=48, b=10),
        yaxis=dict(tickformat=axfmt, title="", gridcolor=GRID_CLR,
                   zeroline=True, zerolinecolor=ZERO_CLR),
        xaxis=dict(title=""),
        plot_bgcolor="rgba(0,0,0,0)",
        bargap=0.25,
        showlegend=False,
    )

    return fig


def _family_bar_fig_multi(family_base, entries, fmt):
    """Multi-year family chart on screen: one bar cluster per bank, one
    color per year, legend along the bottom identifying the years."""
    entries = sorted(entries, key=lambda e: (e[0] is None, e[0] or 0))
    years = [y for y, _ in entries]
    banks = resolve_family_bank_order(entries, num_v, sort_mode)

    if not banks:
        return None

    fig = go.Figure()

    for year, metric in entries:
        excluded = set(chart_excluded_banks(metric))
        series = num_v.loc[metric]

        y_vals, line_widths = [], []
        for b in banks:
            if (b in excluded or b not in series.index
                    or pd.isna(series[b])):
                y_vals.append(None)
                line_widths.append(0)
                continue
            y_vals.append(series[b])
            line_widths.append(3 if b == highlight else 0)

        fig.add_trace(go.Bar(
            name=str(year) if year is not None else display_metric_name(metric),
            x=[short_name(b) for b in banks],
            y=y_vals,
            marker=dict(
                color=year_color(year, years),
                line=dict(color=HILITE, width=line_widths),
            ),
            customdata=[[b, bank_region(b)] for b in banks],
            hovertemplate=("<b>%{customdata[0]}</b> · %{customdata[1]}"
                           "<br>%{y}<extra></extra>"),
            cliponaxis=False,
        ))

    subline_parts = []
    if fmt == "currency":
        subline_parts.append("in $ millions (USD)")

    title_text = display_metric_name(family_base)
    if subline_parts:
        title_text += ("<br><sup style='color:#8a949c'>"
                       + "  ·  ".join(subline_parts) + "</sup>")

    if fmt == "currency":
        axfmt = "$,.0f"
    elif fmt == "percent":
        axfmt = ".1f"
    else:
        axfmt = ",.1f"

    fig.update_layout(
        title=dict(text=title_text, font=dict(size=14)),
        height=330,
        margin=dict(l=10, r=10, t=48, b=10),
        yaxis=dict(tickformat=axfmt, title="", gridcolor=GRID_CLR,
                   zeroline=True, zerolinecolor=ZERO_CLR),
        xaxis=dict(title=""),
        plot_bgcolor="rgba(0,0,0,0)",
        bargap=0.25,
        bargroupgap=0.08,
        barmode="group",
        showlegend=True,
        legend=dict(orientation="h", yanchor="top", y=-0.22,
                    xanchor="center", x=0.5),
    )

    if median_group:
        median = industry_median_for_group(family_base, median_group)
        if median:
            label, value = median
            clr = MEDIAN_GROUP_COLORS.get(median_group, "#5b6770")
            fig.add_hline(
                y=value, line_dash="dash", line_color=clr, line_width=2,
                annotation_text=f"<b>{label}</b>",
                annotation_position="top left",
                annotation_bgcolor="white",
                annotation_bordercolor=clr,
                annotation_borderwidth=1,
                annotation_borderpad=4,
                annotation_font=dict(size=12, color=clr),
            )

    return fig


def _render_family_fig(family_base, entries):
    """Dispatch a family to single-year or multi-year on-screen chart."""
    if len(entries) <= 1:
        metric = entries[0][1]
        return _metric_bar_fig(metric, formats[metric])
    return _family_bar_fig_multi(family_base, entries, formats[entries[0][1]])


with tab_charts:
    topic_pages = [
        (t, ms)
        for t, ms in group_metrics_by_topic(sel_metrics, formats)
        if ms
    ]

    if not topic_pages:
        st.info("No chartable metrics.")
    else:
        page_tabs = st.tabs([
            f"{TOPIC_ICONS.get(t, '📊')} {t} ({len(ms)})"
            for t, ms in topic_pages
        ])

        for page_tab, (topic, page_metrics) in zip(page_tabs, topic_pages):
            with page_tab:
                if topic == TOPIC_MARKET:
                    with st.container(border=True):
                        st.markdown(MARKET_NOTE)
                else:
                    show_reporting_dates()

                cols = st.columns(2)
                families = list(group_into_families(page_metrics))
                shown = 0

                if use_option_1 and topic == TOPIC_PERFORMANCE:
                    # Pair each currency family with the next YoY family in
                    # the SAME row: currency goes left, YoY goes right.
                    for family_base, entries in families:
                        fig = _render_family_fig(family_base, entries)
                        if fig is None:
                            continue
                        is_yoy = "yoy" in family_base.lower()
                        target = cols[1 if is_yoy else 0]
                        target.plotly_chart(fig, use_container_width=True)
                        shown += 1
                else:
                    for family_base, entries in families:
                        fig = _render_family_fig(family_base, entries)
                        if fig is None:
                            continue
                        cols[shown % 2].plotly_chart(
                            fig, use_container_width=True)
                        shown += 1

                if shown == 0:
                    st.info("No numeric values to chart for this topic.")

                show_source_note()


st.divider()

st.markdown("##### Bank legend")

legend_cols = st.columns(3)

for i, b in enumerate(sel_banks):
    legend_cols[i % 3].markdown(f"**{b}** — {full_bank_name(b)}")

